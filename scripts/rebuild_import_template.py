# -*- coding: ascii -*-
"""
rebuild_import_template.py  (ASCII-only strings)
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = Path("public/templates/import-product-variants.xlsx")

TAX_CODES = [
    ("IVA19", "IVA 19%"),
    ("IVA5",  "IVA 5%"),
    ("IVA0",  "IVA 0% (Exento)"),
]
INVENTORY_TYPES = [
    ("REVENTA",     "Producto para reventa"),
    ("SERVICIO",    "Servicio (sin inventario)"),
    ("MANUFACTURA", "Producto fabricado"),
    ("BUNDLE",      "Paquete/combo"),
]
BOOL_VALUES = ["TRUE", "FALSE"]
STD_CODE_TYPES = [
    ("UNSPSC", "Codigo global de producto/servicio"),
    ("EAN",    "Codigo de barras EAN-13"),
    ("GTIN",   "Numero global de articulo comercial"),
    ("PARTNUM","Numero de parte fabricante"),
]

COLUMNS = [
    ("product_name",       "product_name *",     "R"),
    ("category_name",      "category_name",       "O"),
    ("unit_code",          "unit_code",           "O"),
    ("description",        "description",         "O"),
    ("variant_name",       "variant_name",        "O"),
    ("initial_stock",      "initial_stock",       "O"),
    ("unit_cost",          "unit_cost",           "O"),
    ("unit_price",         "unit_price *",        "R"),
    ("tax_code",           "tax_code",            "O"),
    ("price_includes_tax", "price_includes_tax",  "O"),
    ("inventory_type",     "inventory_type",      "O"),
    ("is_active",          "is_active",           "O"),
    ("control_expiration", "control_expiration",  "O"),
    ("is_component",       "is_component",        "O"),
    ("location_code",      "location_code",       "O"),
    ("standard_code",      "standard_code",       "O"),
    ("standard_code_type", "standard_code_type",  "O"),
]

EXAMPLE = {
    "product_name": "Camiseta Basica", "category_name": "Camisetas",
    "unit_code": "UND", "description": "Tela de algodon", "variant_name": "",
    "initial_stock": 100, "unit_cost": 12000, "unit_price": 19000,
    "tax_code": "IVA19", "price_includes_tax": "FALSE", "inventory_type": "REVENTA",
    "is_active": "TRUE", "control_expiration": "FALSE", "is_component": "FALSE",
    "location_code": "PRINCIPAL",
    "standard_code": "", "standard_code_type": "UNSPSC",
}

INST = [
    ("INSTRUCCIONES PLANTILLA IMPORTACION MASIVA", "H"),
    ("", None),
    ("COLUMNAS OBLIGATORIAS (naranja):", "R_H"),
    ("  * product_name  -> Nombre del producto. Si ya existe, se actualiza.", None),
    ("  * unit_price    -> Precio de venta de la variante.", None),
    ("", None),
    ("COLUMNAS OPCIONALES (ambar):", "O_H"),
    ("  * category_name       -> Se crea automaticamente si no existe.", None),
    ("  * unit_code           -> Codigo DIAN. Usa el desplegable (ver hoja unidades_dian).", None),
    ("  * variant_name        -> En blanco = variante Predeterminada.", None),
    ("  * tax_code            -> IVA19 (19%), IVA5 (5%), IVA0 (0%). Ver hoja listas_ref.", None),
    ("  * price_includes_tax  -> TRUE si unit_price ya incluye IVA; FALSE si es precio base.", None),
    ("  * inventory_type      -> REVENTA | SERVICIO | MANUFACTURA | BUNDLE  (default REVENTA).", None),
    ("  * is_active           -> TRUE/FALSE (default TRUE).", None),
    ("  * control_expiration  -> TRUE para activar gestion de lotes y vencimientos.", None),
    ("  * is_component        -> TRUE si es un insumo (no aparece en POS).", None),
    ("  * location_code       -> Nombre exacto de la bodega para el stock inicial.", None),
    ("  * standard_code       -> Codigo UNSPSC/EAN del producto para el XML FE (opcional).", None),
    ("  * standard_code_type  -> Tipo de codigo: UNSPSC | EAN | GTIN | PARTNUM (default UNSPSC).", None),
    ("", None),
    ("NOTAS:", "NOTE"),
    ("  * No modifiques los nombres de las columnas (fila 1).", None),
    ("  * No elimines las hojas productos, unidades_dian o listas_ref.", None),
    ("  * Las filas en blanco se ignoran.", None),
    ("  * Booleanos aceptan: TRUE / FALSE (en mayusculas).", None),
]

def mf(c): return PatternFill(fill_type="solid", fgColor=c)
def tb():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def add_dv(ws, col, sheet, col_ref, last):
    f = "={s}!${c}$2:${c}${l}".format(s=sheet, c=col_ref, l=last)
    dv = DataValidation(type="list", formula1=f, showDropDown=False,
                        showErrorMessage=True, errorTitle="Valor invalido",
                        error="Selecciona de la lista ({})".format(sheet))
    dv.sqref = "{c}2:{c}1001".format(c=col)
    ws.add_data_validation(dv)

def build():
    wb = openpyxl.load_workbook(str(TEMPLATE_PATH))

    # listas_ref
    if "listas_ref" not in wb.sheetnames: wb.create_sheet("listas_ref")
    wr = wb["listas_ref"]
    wr.delete_rows(1, wr.max_row + 1)
    hf = mf("37474F"); hfont = Font(bold=True, color="FFFFFF", size=10)
    for col, items, hdr in [
        (1, TAX_CODES, "tax_code"),
        (4, INVENTORY_TYPES, "inventory_type"),
        (9, STD_CODE_TYPES, "standard_code_type"),
    ]:
        wr.cell(1, col, hdr).fill = hf; wr.cell(1, col).font = hfont
        wr.cell(1, col+1, "Descripcion").fill = hf; wr.cell(1, col+1).font = hfont
        for i, (code, desc) in enumerate(items, 2):
            wr.cell(i, col, code); wr.cell(i, col+1, desc)
    wr.cell(1, 7, "booleanos").fill = hf; wr.cell(1, 7).font = hfont
    for i, v in enumerate(BOOL_VALUES, 2): wr.cell(i, 7, v)
    for c, w in [("A",14),("B",30),("D",16),("E",30),("G",10)]:
        wr.column_dimensions[c].width = w

    # productos
    if "productos" in wb.sheetnames: del wb["productos"]
    ws = wb.create_sheet("productos", 1)
    rfill=mf("E65100"); ofill=mf("F57F17")
    rfont=Font(bold=True,color="FFFFFF",size=9); ofont=Font(bold=False,color="FFFFFF",size=9)
    for ci, (name, label, col) in enumerate(COLUMNS, 1):
        c = ws.cell(1, ci, label)
        c.fill = rfill if col=="R" else ofill
        c.font = rfont if col=="R" else ofont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = tb()
    ws.row_dimensions[1].height = 36
    ef = mf("E3F2FD")
    for ci, (name, _l, _c) in enumerate(COLUMNS, 1):
        c = ws.cell(2, ci, EXAMPLE.get(name,""))
        c.fill = ef; c.border = tb()
        c.alignment = Alignment(horizontal="left", vertical="center")
    for i, w in enumerate([22,18,14,28,18,14,14,14,12,18,16,12,18,14,18,22,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ud_last = wb["unidades_dian"].max_row
    add_dv(ws, "C", "unidades_dian", "A", ud_last)
    add_dv(ws, "I", "listas_ref", "A", len(TAX_CODES)+1)
    add_dv(ws, "K", "listas_ref", "D", len(INVENTORY_TYPES)+1)
    for bc in ["J","L","M","N"]: add_dv(ws, bc, "listas_ref", "G", 3)
    add_dv(ws, "Q", "listas_ref", "I", len(STD_CODE_TYPES)+1)

    # Instrucciones
    if "Instrucciones" in wb.sheetnames: del wb["Instrucciones"]
    wi = wb.create_sheet("Instrucciones", 0)
    wi.column_dimensions["A"].width = 95
    fm = {"H":mf("1A237E"),"R_H":mf("E65100"),"O_H":mf("F57F17"),"NOTE":mf("FFF9C4")}
    fnm = {"H":Font(bold=True,color="FFFFFF",size=14),
           "R_H":Font(bold=True,color="FFFFFF",size=10),
           "O_H":Font(bold=True,color="FFFFFF",size=10),
           "NOTE":Font(bold=True,size=10)}
    for ri, (text, fk) in enumerate(INST, 1):
        c = wi.cell(ri, 1, text)
        c.font = fnm.get(fk, Font(size=10))
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if fk and fk in fm: c.fill = fm[fk]
        wi.row_dimensions[ri].height = 6 if not text else 18

    # order sheets
    desired = ["Instrucciones","productos","unidades_dian","listas_ref"]
    for i, name in enumerate(desired):
        if name in wb.sheetnames:
            cur = wb.sheetnames.index(name)
            if cur != i: wb.move_sheet(name, offset=i-cur)

    wb.save(str(TEMPLATE_PATH))
    print("OK - saved to", TEMPLATE_PATH)
    for dv in ws.data_validations.dataValidation:
        print(" ", dv.sqref, "->", dv.formula1)

if __name__ == "__main__":
    build()