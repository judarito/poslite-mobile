# -*- coding: ascii -*-
"""
rebuild_third_parties_template.py
Generates public/templates/import-third-parties.xlsx
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = Path("public/templates/import-third-parties.xlsx")

# -- reference lists ---------------------------------------------------------
TYPES = [
    ("customer",   "Solo cliente"),
    ("supplier",   "Solo proveedor"),
    ("both",       "Cliente y proveedor"),
    ("employee",   "Empleado"),
]
DOC_TYPES = [
    ("NIT",  "NIT (Empresa)"),
    ("CC",   "Cedula de ciudadania"),
    ("CE",   "Cedula de extranjeria"),
    ("TI",   "Tarjeta de identidad"),
    ("PPN",  "Pasaporte"),
    ("DIE",  "Documento de identidad extranjero"),
    ("RUT",  "RUT"),
]
TAX_REGIMES = [
    ("SIMPLE",     "Regimen simple de tributacion"),
    ("ORDINARIO",  "Regimen ordinario (contribuyente)"),
    ("NO_RESP",    "No responsable de IVA"),
    ("GRAN_CONTRIB","Gran contribuyente"),
]
BOOL_VALUES = ["TRUE", "FALSE"]
CURRENCIES  = ["COP", "USD", "EUR"]

# -- columns -----------------------------------------------------------------
# (internal_name, header_label, colour)  R=required O=optional
COLUMNS = [
    ("legal_name",                  "legal_name *",               "R"),
    ("document_type",               "document_type",              "O"),
    ("document_number",             "document_number",            "O"),
    ("dv",                          "dv",                         "O"),
    ("type",                        "type",                       "O"),
    ("trade_name",                  "trade_name",                 "O"),
    ("phone",                       "phone",                      "O"),
    ("email",                       "email",                      "O"),
    ("fiscal_email",                "fiscal_email",               "O"),
    ("address",                     "address",                    "O"),
    ("city",                        "city",                       "O"),
    ("city_code",                   "city_code",                  "O"),
    ("department",                  "department",                 "O"),
    ("country_code",                "country_code",               "O"),
    ("tax_regime",                  "tax_regime",                 "O"),
    ("is_responsible_for_iva",      "is_responsible_for_iva",     "O"),
    ("obligated_accounting",        "obligated_accounting",       "O"),
    ("default_payment_terms",       "default_payment_terms",      "O"),
    ("max_credit_amount",           "max_credit_amount",          "O"),
    ("default_currency",            "default_currency",           "O"),
    ("is_active",                   "is_active",                  "O"),
]

EXAMPLE = {
    "legal_name":               "Empresa Ejemplo SAS",
    "document_type":            "NIT",
    "document_number":          "900123456",
    "dv":                       "1",
    "type":                     "both",
    "trade_name":               "Ejemplo",
    "phone":                    "3001234567",
    "email":                    "contacto@ejemplo.com",
    "fiscal_email":             "",
    "address":                  "Calle 10 # 5-20",
    "city":                     "Bogota",
    "city_code":                "11001",
    "department":               "Cundinamarca",
    "country_code":             "CO",
    "tax_regime":               "ORDINARIO",
    "is_responsible_for_iva":   "TRUE",
    "obligated_accounting":     "FALSE",
    "default_payment_terms":    30,
    "max_credit_amount":        "",
    "default_currency":         "COP",
    "is_active":                "TRUE",
}

INST = [
    ("INSTRUCCIONES PLANTILLA IMPORTACION DE TERCEROS", "H"),
    ("", None),
    ("COLUMNAS OBLIGATORIAS (naranja):", "R_H"),
    ("  * legal_name  -> Razon social o nombre completo (obligatorio).", None),
    ("", None),
    ("COLUMNAS OPCIONALES (ambar):", "O_H"),
    ("  * document_type         -> Tipo de documento. Usa el desplegable: NIT, CC, CE, etc.", None),
    ("  * document_number       -> Numero de documento (sin DV para NIT).", None),
    ("  * dv                    -> Digito de verificacion (solo NIT).", None),
    ("  * type                  -> customer | supplier | both | employee  (default both).", None),
    ("  * trade_name            -> Nombre comercial o alias.", None),
    ("  * phone                 -> Telefono de contacto.", None),
    ("  * email                 -> Correo electronico.", None),
    ("  * fiscal_email          -> Correo para envio de facturas electronicas.", None),
    ("  * address               -> Direccion (texto libre).", None),
    ("  * city                  -> Ciudad.", None),
    ("  * city_code             -> Codigo DANE del municipio (5 digitos, ej. 11001 = Bogota).", None),
    ("  * department            -> Departamento.", None),
    ("  * country_code          -> Codigo pais (default CO).", None),
    ("  * tax_regime            -> Regimen tributario. Ver desplegable.", None),
    ("  * is_responsible_for_iva -> TRUE/FALSE (default FALSE).", None),
    ("  * obligated_accounting  -> TRUE/FALSE (default FALSE).", None),
    ("  * default_payment_terms -> Dias de plazo de pago (ej. 30).", None),
    ("  * max_credit_amount     -> Limite de credito en pesos.", None),
    ("  * default_currency      -> Moneda: COP, USD, EUR (default COP).", None),
    ("  * is_active             -> TRUE/FALSE (default TRUE).", None),
    ("", None),
    ("NOTAS:", "NOTE"),
    ("  * No modifiques los nombres de las columnas (fila 1).", None),
    ("  * Si document_number ya existe para el tenant, se actualizara el registro.", None),
    ("  * Las filas en blanco se ignoran.", None),
    ("  * Booleanos aceptan: TRUE / FALSE (en mayusculas).", None),
]

def mf(c): return PatternFill(fill_type="solid", fgColor=c)
def tb():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def add_dv(ws, col, sheet, col_ref, last):
    f = "={s}!${c}$2:${c}${l}".format(s=sheet, c=col_ref, l=last)
    dv = DataValidation(type="list", formula1=f, showDropDown=False,
                        showErrorMessage=True, errorTitle="Valor invalido",
                        error="Selecciona de la lista ({})".format(sheet))
    dv.sqref = "{c}2:{c}1001".format(c=col)
    ws.add_data_validation(dv)

def add_dv_inline(ws, col, values):
    joined = ",".join(values)
    dv = DataValidation(type="list", formula1='"{}"'.format(joined),
                        showDropDown=False, showErrorMessage=True,
                        errorTitle="Valor invalido",
                        error="Elige: {}".format(joined))
    dv.sqref = "{c}2:{c}1001".format(c=col)
    ws.add_data_validation(dv)

def build():
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # ---- listas_ref --------------------------------------------------------
    ws_ref = wb.active
    ws_ref.title = "listas_ref"
    hf = mf("37474F"); hfont = Font(bold=True, color="FFFFFF", size=10)

    for col, items, hdr in [
        (1, TYPES,       "type"),
        (4, DOC_TYPES,   "document_type"),
        (7, TAX_REGIMES, "tax_regime"),
        (10, [(v,"") for v in BOOL_VALUES], "booleanos"),
        (12, [(v,"") for v in CURRENCIES], "monedas"),
    ]:
        ws_ref.cell(1, col, hdr).fill = hf
        ws_ref.cell(1, col).font = hfont
        if items and items[0][1]:
            ws_ref.cell(1, col+1, "Descripcion").fill = hf
            ws_ref.cell(1, col+1).font = hfont
        for i, (code, desc) in enumerate(items, 2):
            ws_ref.cell(i, col, code)
            if desc: ws_ref.cell(i, col+1, desc)

    for c, w in [("A",14),("B",28),("D",10),("E",32),("G",14),("H",36),("J",10),("L",8)]:
        ws_ref.column_dimensions[c].width = w

    # ---- terceros sheet ----------------------------------------------------
    ws = wb.create_sheet("terceros", 0)
    rfill=mf("E65100"); ofill=mf("F57F17")
    rfont=Font(bold=True,color="FFFFFF",size=9); ofont=Font(bold=False,color="FFFFFF",size=9)

    for ci, (name, label, col) in enumerate(COLUMNS, 1):
        c = ws.cell(1, ci, label)
        c.fill = rfill if col=="R" else ofill
        c.font = rfont if col=="R" else ofont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = tb()
    ws.row_dimensions[1].height = 36

    ef = mf("E3F2FD")
    for ci, (name, _l, _c) in enumerate(COLUMNS, 1):
        c = ws.cell(2, ci, EXAMPLE.get(name,""))
        c.fill = ef; c.border = tb()
        c.alignment = Alignment(horizontal="left", vertical="center")

    widths = [26,14,18,6,12,18,14,26,26,28,16,14,18,12,14,18,18,18,18,16,12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # data validations
    # Col B = document_type (col 4 in listas_ref -> D)
    add_dv(ws, "B", "listas_ref", "D", len(DOC_TYPES)+1)
    # Col E = type (col 1 in listas_ref -> A)
    add_dv(ws, "E", "listas_ref", "A", len(TYPES)+1)
    # Col O = tax_regime (col 7 -> G)  [shifted by city_code in col 12]
    add_dv(ws, "O", "listas_ref", "G", len(TAX_REGIMES)+1)
    # Col T = default_currency
    add_dv(ws, "T", "listas_ref", "L", len(CURRENCIES)+1)
    # Booleans: P(is_responsible_for_iva), Q(obligated_accounting), U(is_active) -> col 10 (J)
    for bc in ["P", "Q", "U"]:
        add_dv(ws, bc, "listas_ref", "J", 3)

    # ---- Instrucciones sheet -----------------------------------------------
    wi = wb.create_sheet("Instrucciones", 0)
    wi.column_dimensions["A"].width = 95
    fm = {"H":mf("1A237E"),"R_H":mf("E65100"),"O_H":mf("F57F17"),"NOTE":mf("FFF9C4")}
    fnm = {"H":Font(bold=True,color="FFFFFF",size=14),
           "R_H":Font(bold=True,color="FFFFFF",size=10),
           "O_H":Font(bold=True,color="FFFFFF",size=10),
           "NOTE":Font(bold=True,size=10)}
    for ri, (text, fk) in enumerate(INST, 1):
        c = wi.cell(ri, 1, text)
        c.font = fnm.get(fk, Font(size=10))
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if fk and fk in fm: c.fill = fm[fk]
        wi.row_dimensions[ri].height = 6 if not text else 18

    # sheet order: Instrucciones, terceros, listas_ref
    desired = ["Instrucciones","terceros","listas_ref"]
    for i, name in enumerate(desired):
        if name in wb.sheetnames:
            cur = wb.sheetnames.index(name)
            if cur != i: wb.move_sheet(name, offset=i-cur)

    wb.save(str(TEMPLATE_PATH))
    print("OK:", TEMPLATE_PATH)
    print("Sheets:", wb.sheetnames)
    for dv in ws.data_validations.dataValidation:
        print(" ", dv.sqref, "->", dv.formula1)

if __name__ == "__main__":
    build()